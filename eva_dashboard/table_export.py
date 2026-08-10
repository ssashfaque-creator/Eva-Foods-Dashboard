"""Export chatbot sales / party tables to Excel or PDF (readable layout)."""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from eva_dashboard.party_analytics import analyze_parties, list_clients
from eva_dashboard.sales_query import query_sales

BRAND = colors.Color(0.09, 0.29, 0.22)
HEADER_BG = colors.Color(0.09, 0.29, 0.22)
HEADER_FG = colors.white
ROW_ALT = colors.Color(0.94, 0.96, 0.95)
LINE = colors.Color(0.75, 0.82, 0.78)
SUBTOTAL_BG = colors.Color(0.88, 0.93, 0.90)
TOTAL_BG = colors.Color(0.78, 0.88, 0.84)


def _safe_filename(title: str, ext: str) -> str:
    stem = re.sub(r"[^\w\-]+", "_", (title or "eva_table").strip())[:60].strip("_")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem or 'eva_table'}_{stamp}.{ext}"


def matrix_to_records(
    matrix: dict[str, Any],
) -> tuple[list[str], list[list[Any]], int]:
    """Flatten a sales matrix (flat or hierarchical) to header + data rows.

    Returns ``(headers, data, dim_count)`` where ``dim_count`` is the number of
    leading row-dimension columns.
    """
    if not matrix:
        return [], [], 0
    row_headers = list(
        matrix.get("row_headers")
        or (
            [matrix["row_dimension"]]
            if matrix.get("row_dimension")
            else ["Row"]
        )
    )
    value_cols = [str(c) for c in (matrix.get("columns") or [])]
    headers = [str(h) for h in row_headers] + value_cols
    data: list[list[Any]] = []
    for r in matrix.get("rows") or []:
        cells: list[Any] = []
        for h in row_headers:
            cells.append(r.get(h, ""))
        for c in value_cols:
            cells.append(r.get(c, ""))
        data.append(cells)
    return headers, data, len(row_headers)


def rows_to_records(
    rows: list[dict[str, Any]],
    columns: list[str] | None = None,
) -> tuple[list[str], list[list[Any]]]:
    if not rows:
        return list(columns or []), []
    cols = list(columns) if columns else list(rows[0].keys())
    data = [[r.get(c, "") for c in cols] for r in rows]
    return cols, data


def _labelize(name: str) -> str:
    mapping = {
        "business_unit": "Business Unit",
        "packing_category": "Packing",
        "oil_type": "Oil Type",
        "client_type": "Client Type",
        "city": "City",
        "zone": "Zone",
        "party": "Client",
        "product": "Product",
        "mt": "MT",
        "volume_mt": "Volume (MT)",
        "ams_mt": "AMS (MT)",
        "client": "Client",
        "city_filter": "City-Filter",
    }
    if name in mapping:
        return mapping[name]
    if name.startswith("AMS") or name.endswith("%") or re.match(r"^[A-Z][a-z]{2} \d{4}$", name):
        return name
    return name.replace("_", " ").title()


def dataframe_from_records(headers: list[str], data: list[list[Any]]) -> pd.DataFrame:
    nice = [_labelize(h) for h in headers]
    frame = pd.DataFrame(data, columns=nice)
    return frame


def build_excel_bytes(
    *,
    title: str,
    headers: list[str],
    data: list[list[Any]],
    subtitle: str | None = None,
) -> bytes:
    frame = dataframe_from_records(headers, data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet = "Table"
        # Title rows above the table
        start_row = 3 if subtitle else 2
        frame.to_excel(writer, sheet_name=sheet, index=False, startrow=start_row)
        ws = writer.sheets[sheet]
        ws["A1"] = title or "Eva Foods table"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="17334A")
        if subtitle:
            ws["A2"] = subtitle
            ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")

        header_fill = PatternFill("solid", fgColor="174B38")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        alt_fill = PatternFill("solid", fgColor="F0F5F2")
        thin = Border(
            left=Side(style="thin", color="C0CFC8"),
            right=Side(style="thin", color="C0CFC8"),
            top=Side(style="thin", color="C0CFC8"),
            bottom=Side(style="thin", color="C0CFC8"),
        )
        header_row = start_row + 1
        for col_i in range(1, len(frame.columns) + 1):
            cell = ws.cell(header_row, col_i)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        for r_i, row in enumerate(
            ws.iter_rows(
                min_row=header_row + 1,
                max_row=header_row + len(frame),
                min_col=1,
                max_col=len(frame.columns),
            ),
            start=0,
        ):
            for c_i, cell in enumerate(row):
                cell.border = thin
                cell.alignment = Alignment(
                    horizontal="right" if c_i > 0 else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if r_i % 2 == 1:
                    cell.fill = alt_fill
                # Emphasize Total / subtotal rows
                first = str(frame.iloc[r_i, 0] if len(frame.columns) else "")
                if "total" in first.lower():
                    cell.font = Font(name="Calibri", bold=True)

        ws.row_dimensions[header_row].height = 22
        for col_i, col_name in enumerate(frame.columns, start=1):
            series = frame.iloc[:, col_i - 1].astype(str)
            width = min(
                42,
                max(
                    10,
                    len(str(col_name)) + 2,
                    int(series.map(len).max() if len(series) else 8) + 2,
                ),
            )
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.freeze_panes = f"A{header_row + 1}"
    return buf.getvalue()


def build_pdf_bytes(
    *,
    title: str,
    headers: list[str],
    data: list[list[Any]],
    subtitle: str | None = None,
    dim_count: int | None = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title or "Eva Foods table",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "EvaExportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=14,
        textColor=BRAND,
        spaceAfter=2 * mm,
        alignment=TA_LEFT,
    )
    meta_style = ParagraphStyle(
        "EvaExportMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.Color(0.35, 0.4, 0.38),
        spaceAfter=4 * mm,
    )
    cell_style = ParagraphStyle(
        "EvaExportCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_LEFT,
    )
    num_style = ParagraphStyle(
        "EvaExportNum",
        parent=cell_style,
        alignment=TA_CENTER,
    )
    head_style = ParagraphStyle(
        "EvaExportHead",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=HEADER_FG,
        alignment=TA_CENTER,
    )

    nice = [_labelize(h) for h in headers]
    if dim_count is None:
        dim_count = 0
        for h in headers:
            if re.match(
                r"^(AMS|Total|MT|Volume|YoY|%|Share)|^[A-Z][a-z]{2} \d{4}$",
                str(h),
            ) or str(h).endswith("%"):
                break
            dim_count += 1
    dim_count = max(1, int(dim_count or 1))

    table_data: list[list[Any]] = [
        [Paragraph(escape(str(h)), head_style) for h in nice]
    ]
    for row in data:
        out_row = []
        for i, val in enumerate(row):
            text = "" if val is None else str(val)
            style = cell_style if i < dim_count else num_style
            out_row.append(Paragraph(escape(text), style))
        table_data.append(out_row)

    col_count = max(1, len(nice))
    page_w = landscape(A4)[0] - 20 * mm
    # Give dimension columns more width
    widths: list[float] = []
    for i in range(col_count):
        if i < dim_count:
            widths.append(page_w * (0.22 if dim_count == 1 else 0.16))
        else:
            widths.append(page_w * 0.78 / max(1, col_count - dim_count))
    # Normalize to page width
    total_w = sum(widths) or 1
    widths = [w * page_w / total_w for w in widths]

    table = Table(table_data, colWidths=widths, repeatRows=1)
    style_cmds: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (dim_count, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    for r_i in range(1, len(table_data)):
        if r_i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, r_i), (-1, r_i), ROW_ALT))
        first = str(data[r_i - 1][0] if data and data[r_i - 1] else "").lower()
        if "total" in first:
            style_cmds.append(("BACKGROUND", (0, r_i), (-1, r_i), TOTAL_BG))
            style_cmds.append(("FONTNAME", (0, r_i), (-1, r_i), "Helvetica-Bold"))
        elif "subtotal" in first or first.endswith(" total"):
            style_cmds.append(("BACKGROUND", (0, r_i), (-1, r_i), SUBTOTAL_BG))
    table.setStyle(TableStyle(style_cmds))

    story = [Paragraph(escape(title or "Eva Foods table"), title_style)]
    if subtitle:
        story.append(Paragraph(escape(subtitle), meta_style))
    story.append(Spacer(1, 2 * mm))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def query_sales_from_table_spec(spec: dict[str, Any]) -> dict[str, Any]:
    """Rebuild a sales result from a stored table_spec."""
    filters = dict(spec.get("filters") or {})
    period = spec.get("period") or {}
    return query_sales(
        period=spec.get("period_phrase"),
        date_from=period.get("date_from"),
        date_to=period.get("date_to"),
        city=filters.get("city"),
        zone=filters.get("zone"),
        business_unit=filters.get("business_unit"),
        business_units=list(spec.get("business_units") or []) or None,
        oil_type=filters.get("oil_type"),
        packing_category=filters.get("packing_category"),
        client_type=filters.get("client_type"),
        party=filters.get("party"),
        columns=spec.get("column_dimension") or "client_type",
        months_back=int(spec.get("months_back") or 6),
        row_dimension=spec.get("row_dimension"),
        row_groups=list(spec.get("row_groups") or []) or None,
        excludes=spec.get("excludes"),
        lock_columns=True,
        mode="matrix",
        compare=spec.get("compare"),
        prior_spec=None,
    )


def export_payload_from_followup(
    meta: dict[str, Any] | None,
) -> dict[str, Any] | None:
    """Build exportable headers/data from Reply follow-up meta."""
    if not meta:
        return None
    # Prefer the snapshot captured when the answer was built (no re-query)
    snap = meta.get("export")
    if isinstance(snap, dict) and snap.get("headers") is not None:
        return {
            "ok": True,
            "title": snap.get("title") or "Eva Foods table",
            "subtitle": snap.get("subtitle"),
            "headers": list(snap.get("headers") or []),
            "data": list(snap.get("data") or []),
            "dim_count": int(snap.get("dim_count") or 1),
            "filename_stem": snap.get("filename_stem") or "eva_table",
        }
    table_spec = meta.get("table_spec")
    party_spec = meta.get("party_spec")
    if table_spec:
        result = query_sales_from_table_spec(table_spec)
        if not result.get("ok"):
            return {"ok": False, "error": result.get("error") or "Could not rebuild table"}
        headers, data, dim_count = matrix_to_records(result.get("matrix") or {})
        title = "Eva Foods sales table"
        period = (result.get("period") or {}).get("label") or ""
        filters = result.get("filters") or {}
        bits = [period] if period else []
        for key in ("city", "zone", "client_type", "business_unit"):
            if filters.get(key):
                bits.append(f"{key}={filters[key]}")
        return {
            "ok": True,
            "title": title,
            "subtitle": " · ".join(bits),
            "headers": headers,
            "data": data,
            "dim_count": dim_count,
            "filename_stem": "eva_sales_table",
        }
    if party_spec:
        kind = party_spec.get("kind") or "list_clients"
        filters = dict(party_spec.get("filters") or {})
        period = party_spec.get("period") or {}
        if kind == "analyze_parties":
            result = analyze_parties(
                period=party_spec.get("period_phrase"),
                date_from=period.get("date_from"),
                date_to=period.get("date_to"),
                city=filters.get("city"),
                zone=filters.get("zone"),
                client_type=filters.get("client_type"),
                business_unit=filters.get("business_unit"),
                oil_type=filters.get("oil_type"),
                packing_category=filters.get("packing_category"),
                metric=party_spec.get("metric") or "volume",
                limit=int(party_spec.get("limit") or 100),
                group_by=party_spec.get("group_by") or "party",
            )
            rows = list(result.get("parties") or [])
        else:
            result = list_clients(
                city=filters.get("city"),
                zone=filters.get("zone"),
                client_type=filters.get("client_type"),
                business_unit=filters.get("business_unit"),
                period=party_spec.get("period_phrase"),
                date_from=period.get("date_from"),
                date_to=period.get("date_to"),
                limit=int(party_spec.get("limit") or 200),
                include_zero=bool(party_spec.get("include_zero")),
            )
            rows = list(result.get("clients") or [])
        if not result.get("ok"):
            return {"ok": False, "error": result.get("error") or "Could not rebuild table"}
        prefer = [
            "party",
            "client",
            "client_type",
            "zone",
            "city_filter",
            "city",
            "mt",
            "volume_mt",
            "ams_mt",
            "score",
            "yoy_pct",
            "pct_vs_ams",
        ]
        cols = [c for c in prefer if rows and c in rows[0]]
        if not cols and rows:
            cols = list(rows[0].keys())
        headers, data = rows_to_records(rows, cols)
        return {
            "ok": True,
            "title": "Eva Foods clients / parties",
            "subtitle": (period.get("label") or "") or None,
            "headers": headers,
            "data": data,
            "dim_count": 1,
            "filename_stem": "eva_party_table",
        }
    return None


def export_excel_from_followup(meta: dict[str, Any] | None) -> tuple[bytes, str] | None:
    payload = export_payload_from_followup(meta)
    if not payload or not payload.get("ok"):
        return None
    blob = build_excel_bytes(
        title=str(payload.get("title") or "Eva Foods"),
        subtitle=payload.get("subtitle"),
        headers=list(payload.get("headers") or []),
        data=list(payload.get("data") or []),
    )
    return blob, _safe_filename(str(payload.get("filename_stem") or "eva_table"), "xlsx")


def export_pdf_from_followup(meta: dict[str, Any] | None) -> tuple[bytes, str] | None:
    payload = export_payload_from_followup(meta)
    if not payload or not payload.get("ok"):
        return None
    blob = build_pdf_bytes(
        title=str(payload.get("title") or "Eva Foods"),
        subtitle=payload.get("subtitle"),
        headers=list(payload.get("headers") or []),
        data=list(payload.get("data") or []),
        dim_count=int(payload.get("dim_count") or 1),
    )
    return blob, _safe_filename(str(payload.get("filename_stem") or "eva_table"), "pdf")
